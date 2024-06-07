import openpyxl

from api.models import Category, PathologyTest

def migrate():
    # Load the Excel file
    file_path = 'tests.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Helper function to get cell value
    def get_cell_value(row, col):
        cell = sheet.cell(row=row, column=col)
        return cell.value

    # Iterate over each row in the worksheet and create PathologyTest entries
    for row in range(2, sheet.max_row + 1):  # assuming the first row is the header
        name = get_cell_value(row, 1)
        description = get_cell_value(row, 2)
        test_type = get_cell_value(row, 3)
        category_name = get_cell_value(row, 4)
        regular_price = get_cell_value(row, 5) or 0
        price = get_cell_value(row, 6) or 0
        fasting = get_cell_value(row, 7) or "No"
        gender = get_cell_value(row, 8) or "Male, Female"
        age = get_cell_value(row, 9) or "5-99"
        report_time = get_cell_value(row, 10) or "Reports within 24 hrs"
        
        
        if(price == None or not category_name or not test_type or not description):            
            continue
        else:
            print(name, description, test_type, category_name, regular_price, price, fasting, gender, age, report_time)
            
        # Get or create the category instance (assuming it exists)
        category, created = Category.objects.get_or_create(category_name=category_name, defaults={'description': ''})
        
        # Create the PathologyTest instance
        pathology_test = PathologyTest.objects.create(
            name=name,
            description=description,
            test_type=test_type,
            category=category,
            regular_price=regular_price,
            price=price,
            fasting=fasting,
            gender=gender,
            age=age,
            report_time=report_time,
            pathology_id = 4
        )

        pathology_test.save()
